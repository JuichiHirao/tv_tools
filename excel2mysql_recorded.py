import openpyxl
import sys
import traceback
import re
from db import TvRecordedDao, TvProgramDao, TvDiskDao
from common import RecordedTool
from data import TvRecordedData
from datetime import datetime
from logging import getLogger, DEBUG, Formatter, FileHandler, StreamHandler

str_date = datetime.now().strftime('%Y%m%d')
logger = getLogger("logger")
logger.setLevel(DEBUG)

log_filename = 'log\\tv_recorded_{}.log'.format(str_date)
handler_file = FileHandler(filename=log_filename, encoding='utf-8')
handler_file.setFormatter(Formatter("%(asctime)s %(levelname)4s %(message)s"))
logger.addHandler(handler_file)
handler_stream = StreamHandler(sys.stdout)
handler_stream.setFormatter(Formatter("%(asctime)s %(levelname)8s %(message)s"))
logger.addHandler(handler_stream)

class TvContentsRegister:

    def __init__(self, range_disk_no: str = '', is_check: bool = True):
        self.recorded_dao = TvRecordedDao()
        self.program_dao = TvProgramDao()
        self.excel_file = 'C:\\Users\\JuichiHirao\\Dropbox\\jhdata\\Interest\\BD番組録画.xlsx'
        self.recorded_tool = RecordedTool(logger)
        self.disk_dao = TvDiskDao()
        # self.base_dir = 'F:\\TVREC'

        self.disk_no_start = 0
        self.disk_no_end = 9999
        if len(range_disk_no) > 0:
            disk_no_list = re.split(',', range_disk_no)
            if len(disk_no_list) == 1:
                self.disk_no_start = int(disk_no_list[0])
                self.disk_no_end = self.disk_no_start
            elif len(disk_no_list) == 2:
                self.disk_no_start = int(disk_no_list[0])
                self.disk_no_end = int(disk_no_list[1])
            else:
                logger.warning(f'range_disk_no {range_disk_no} is invalid')
        self.range_disk_no = range_disk_no
        logger.info(f'range_disk_no start [{self.disk_no_start}] end [{self.disk_no_end}]')
        self.is_check = is_check

    def __get_channel_info(self, program_id_str: str = ''):

        # 4K
        if len(program_id_str) == 7 and program_id_str[0] == '4':
            channel_no = int(program_id_str[0:4])
            channel_seq = int(program_id_str[4:])
        # 663 piggo(1000超え)
        elif len(program_id_str) == 7 and program_id_str[0:3] == '663':
            channel_no = int(program_id_str[0:3])
            channel_seq = int(program_id_str[3:])
        else:
            try:
                channel_no = int(program_id_str[0:3])
                channel_seq = int(program_id_str[3:])
            except ValueError as verr:
                logger.error(f'channel_no/seq [{program_id_str}] {verr}')
                channel_no = 0
                channel_seq = 0

        return channel_no, channel_seq

    def __get_minute(self, time_str):

        if time_str is None or len(str(time_str)) <= 0:
            return 0

        time_list = str(time_str).split(':')
        minute = 0
        if len(time_list) == 2:
            minute = (int(time_list[0]) * 60) + int(time_list[1])

        return minute

    def __get_on_air_datetime(self, date_value, time_value):

        on_air_date_str = None

        try:
            if date_value is not None:
                on_air_date_str = datetime.strftime(date_value, '%Y/%m/%d')
        except ValueError as vex:
            logger.error(traceback.print_exc())
            logger.error(f'except on_air_dateの変換中にエラー発生 date[{date_value}] time[{time_value}] {vex}')

        is_time_flag = True
        if time_value is not None:
            time_list = time_value.split(':')
            if on_air_date_str is None:
                logger.info(f'on_air_date_strがNoneのため、SKIP date[{date_value}] time[{time_value}]')
                return None, False
            else:
                if len(time_list) == 2:
                    on_air_datetime = datetime.strptime('{} {}'.format(on_air_date_str, time_value), '%Y/%m/%d %H:%M')
                else:
                    on_air_datetime = datetime.strptime(on_air_date_str, '%Y/%m/%d')
            # print('on_air_datetime {} '.format(on_air_datetime))
        else:
            on_air_datetime = datetime.strptime(on_air_date_str, '%Y/%m/%d')
            is_time_flag = False
            # print('on_air_datetime {} '.format(on_air_datetime))

        return on_air_datetime, is_time_flag

    def export(self, sheet_name: str = ''):
        """
        「TV録画2」のシートの主要な列以外の列に何が入っているかの確認用
        :return:
        """

        if len(sheet_name) <= 0:
            return

        wb = openpyxl.load_workbook(self.excel_file)
        sheet = wb[sheet_name]
        max_row = sheet.max_row + 1
        # rows = ws[f'A2':'J{max_row}']

        recorded_list = self.recorded_dao.get_where_list()
        program_list = self.program_dao.get_where_list()

        before_disk_label = ''

        col_no_disk_label = 1
        col_no_seq_no = 2
        col_no_rip_status = 3
        col_no_on_air_date = 4
        col_no_program_id = 6
        col_no_on_air_time = 8
        col_no_duration = 9
        col_no_detail = 10

        for row_idx in range(2, max_row):

            disk_label = sheet.cell(row=row_idx, column=col_no_disk_label).value
            seq_no = sheet.cell(row=row_idx, column=col_no_seq_no).value
            rip_status = sheet.cell(row=row_idx, column=col_no_rip_status).value
            if disk_label == 'Disk No':
                continue

            if seq_no is None or seq_no <= 0:
                program_id = sheet.cell(row=row_idx, column=col_no_program_id).value
                channel_no, channel_seq = self.__get_channel_info(str(program_id).zfill(6))
                match_program_list = list(filter(lambda program_data:
                                                 program_data.channelNo == channel_no
                                                 and program_data.channelSeq == channel_seq, program_list))

                program_name = ''
                if len(match_program_list) == 1:
                    program_name = match_program_list[0].name
                elif len(match_program_list) > 1:
                    program_name = f'MANY({len(match_program_list)}) {match_program_list[0].name}'
                on_air_date = sheet.cell(row=row_idx, column=col_no_on_air_date).value
                logger.info(f'[{row_idx}] diskNo ( seqNo & onAirDate ) がないためSKIP [{on_air_date}] 【{program_name}】')
                continue

            disk_no = self.recorded_tool.get_disk_no_label(disk_label, self.disk_dao)

            if disk_no is None or disk_no <= 0:
                disk_no = self.recorded_tool.get_disk_no_label(before_disk_label, self.disk_dao)
                disk_label = before_disk_label
            # if disk_no == 2778:
            #     pass

            if self.disk_no_start > disk_no or self.disk_no_end < disk_no:
                before_disk_label = disk_label
                continue

            tv_data = TvRecordedData()

            program_id = sheet.cell(row=row_idx, column=col_no_program_id).value
            tv_data.channelNo, tv_data.channelSeq = self.__get_channel_info(str(program_id).zfill(6))

            tv_data.diskNo = disk_no
            tv_data.diskLabel = disk_label
            tv_data.seqNo = str(seq_no)
            tv_data.ripStatus = tv_data.get_rip_status(rip_status, '')
            on_air_date = sheet.cell(row=row_idx, column=col_no_on_air_date).value
            tv_data.timeStr = sheet.cell(row=row_idx, column=col_no_on_air_time).value
            tv_data.onAirDate, tv_data.timeFlag = self.__get_on_air_datetime(on_air_date, tv_data.timeStr)
            duration = sheet.cell(row=row_idx, column=col_no_duration).value
            tv_data.minute = self.__get_minute(duration)
            tv_data.detail = sheet.cell(row=row_idx, column=col_no_detail).value
            before_disk_label = tv_data.diskLabel

            tv_data.source = 'excel'

            filter_list = list(filter(lambda recorded_data:
                                      recorded_data.diskNo == tv_data.diskNo
                                      and recorded_data.seqNo == tv_data.seqNo, recorded_list))
            match_program_list = list(filter(lambda program_data:
                                      program_data.channelNo == tv_data.channelNo
                                      and program_data.channelSeq == tv_data.channelSeq, program_list))

            if len(filter_list) == 1:
                is_equal, remark = tv_data.get_update_column(filter_list[0])
                if is_equal is False:
                    tv_data.id = filter_list[0].id
                    tv_data.remark = remark
                    logger.info(f'update target {tv_data.diskNo} {tv_data.seqNo} [{remark}]')
                    if self.is_check is False:
                        self.recorded_dao.update_all(tv_data)
                # else:
                #     print('same data {} {}'.format(tv_data.diskNo, tv_data.seqNo))
            else:
                program_name = ''
                if len(match_program_list) == 1:
                    program_name = match_program_list[0].name
                elif len(match_program_list) > 1:
                    program_name = f'MANY({len(match_program_list)}) {match_program_list[0].name}'
                logger.info(f'register target {tv_data.diskNo} {tv_data.seqNo}')
                logger.info(f'  {tv_data.onAirDate} {tv_data.minute}分 [{tv_data.channelNo}:{tv_data.channelSeq}] 【{program_name}】')
                if self.is_check is False:
                    self.recorded_dao.export(tv_data)

        return

if __name__ == '__main__':
    # range_disk_no = '2780,2789'
    range_disk_no = '2000,2789'
    tv_contents_register = TvContentsRegister(range_disk_no)
    # tv_contents_register = TvContentsRegister(range_disk_no, False)
    # tv_contents_register.export()
    # tv_contents_register.export2('TV録画2')
    # tv_contents_register.export2('0001-1114')
    tv_contents_register.export('ZIP')
    # tv_contents_register.export2('2030')
