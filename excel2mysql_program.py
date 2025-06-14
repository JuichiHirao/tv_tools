import openpyxl
import sys
from db import TvProgramDao
from data import TvProgramData
from datetime import datetime
from logging import getLogger, DEBUG, Formatter, FileHandler, StreamHandler

str_date = datetime.now().strftime('%Y%m%d')
logger = getLogger("logger")
logger.setLevel(DEBUG)

log_filename = 'log\\tv_program_{}.log'.format(str_date)
handler_file = FileHandler(filename=log_filename, encoding='utf-8')
handler_file.setFormatter(Formatter("%(asctime)s %(levelname)4s %(message)s"))
logger.addHandler(handler_file)
handler_stream = StreamHandler(sys.stdout)
handler_stream.setFormatter(Formatter("%(asctime)s %(levelname)8s %(message)s"))
logger.addHandler(handler_stream)


class TvProgramRegister:

    def __init__(self):
        self.program_dao = TvProgramDao()
        self.excel_file = 'C:\\Users\\JuichiHirao\\Dropbox\\jhdata\\Interest\\BD番組録画.xlsx'

        # self.is_clear = True
        self.is_clear = False

        # self.is_check = True
        self.is_check = False

    def export(self):

        if self.is_clear:
            self.program_dao.clear_table()

        col_no_number = 1
        col_no_channel_name = 2 # CHANNELシートのVLOOKUP
        col_no_name = 3
        col_no_short_name = 4
        col_no_start_date = 5
        col_no_end_date = 6
        col_no_category = 7
        col_no_detail = 8
        wb = openpyxl.load_workbook(self.excel_file)
        sheet = wb['番組名']
        max_row = sheet.max_row + 1
        logger.info(f'max_row {max_row}')
        # rows = sheet['A1':'G4000']
        for row_idx in range(1, max_row):
            program_number = sheet.cell(row=row_idx, column=col_no_number).value
            if program_number is None or type(program_number) is not int:
                logger.warning(f'cell_value is None or not int {row_idx} {program_number}')
                continue

            is_update = False
            program_data = TvProgramData()
            program_data.channelNo = int(program_number // 1000)
            program_data.channelSeq = program_number % 1000

            if program_data.channelNo == 663 and program_data.channelSeq >= 925:
                # idx_idx = 0
                continue

            exist_data = self.program_dao.get_data(program_data.channelNo, program_data.channelSeq)

            # UPDATE対象かの確認
            update_column_list = []
            program_data.name = sheet.cell(row=row_idx, column=col_no_name).value
            if program_data.name is None or len(program_data.name.strip()) <= 0:
                continue

            if exist_data is not None and exist_data.name != program_data.name:
                update_column_list.append(f'name {program_data.name} <- {exist_data.name}')
                is_update = True

            program_data.shortName = sheet.cell(row=row_idx, column=col_no_short_name).value
            if exist_data is not None and exist_data.shortName != program_data.shortName:
                update_column_list.append(f'short_name {program_data.shortName} <- {exist_data.shortName}')
                is_update = True

            program_data.startDate = sheet.cell(row=row_idx, column=col_no_start_date).value
            if exist_data is not None and exist_data.startDate != program_data.startDate:
                update_column_list.append(f'start_date {program_data.startDate} <- {exist_data.startDate}')
                is_update = True

            program_data.endDate = sheet.cell(row=row_idx, column=col_no_end_date).value
            if exist_data is not None and exist_data.endDate != program_data.endDate:
                update_column_list.append(f'end_date {program_data.endDate} <- {exist_data.endDate}')
                is_update = True

            program_data.set_date(program_data.startDate, program_data.endDate)

            program_data.category = sheet.cell(row=row_idx, column=col_no_category).value
            if exist_data is not None and exist_data.category != program_data.category:
                if program_data.category is not None and len(program_data.category) > 0:
                    update_column_list.append(f'category {program_data.category} <- {exist_data.category}')
                    is_update = True

            program_data.detail = sheet.cell(row=row_idx, column=col_no_detail).value
            if exist_data is not None and exist_data.detail != program_data.detail:
                update_column_list.append(f'detail {program_data.detail} <- {exist_data.detail}')
                is_update = True

            # if is_update:
            #     logger.info(f'update_column_list {update_column_list}')
            if exist_data is not None:
                if is_update:
                    logger.info(f'exist update target program_data {program_data.channelNo} {program_data.channelSeq} {update_column_list}')
                    self.program_dao.update(program_data, exist_data.id)
                # else:
                #     logger.info(f'same program_data no update {program_data.channelNo} {program_data.channelSeq}')
            else:
                logger.info(f'not exist register target program_data {program_data.channelNo} {program_data.channelSeq} {program_data.name}')
                self.program_dao.export(program_data)

            # if row_idx > 10:
            #     break

        return


if __name__ == '__main__':
    tv_program_register = TvProgramRegister()
    tv_program_register.export()

