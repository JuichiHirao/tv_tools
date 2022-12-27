import openpyxl
import sys
from db import TvRecordedDao
from db import TvProgramDao
from data import TvRecordedData
from datetime import datetime


class TvContentsRegister:

    def __init__(self):
        self.recorded_dao = TvRecordedDao()
        self.program_dao = TvProgramDao()
        self.excel_file = 'C:\\Users\\JuichiHirao\\Dropbox\\jhdata\\Interest\\BD番組録画.xlsx'
        # self.base_dir = 'F:\\TVREC'

        self.is_check = True
        # self.is_check = False

    def __check_program_id(self, target_idx, row):
        # H829 不明なので、OK
        # H2129 不明なので、OK
        program_id = -1
        cell_value = row[target_idx].value
        if cell_value is None:
            print('program id None {} {}'.format(row[target_idx], cell_value))
        elif type(cell_value) is not int:
            print('program id not int {} {} {}'.format(target_idx, row[target_idx], cell_value))
        else:
            program_id = int(cell_value)

        return program_id

    def __get_minute(self, time_str):

        if time_str is None or len(str(time_str)) <= 0:
            return 0

        time_list = str(time_str).split(':')
        minute = 0
        if len(time_list) == 2:
            minute = (int(time_list[0]) * 60) + int(time_list[1])

        return minute

    def __get_cell_data(self, target_idx, row):
        cell_value = row[target_idx].value
        if cell_value is not None and type(cell_value) is not str:
            print('{} {} {} not type str'.format(target_idx, row[target_idx], cell_value))
        elif cell_value is not None and len(cell_value) > 0:
            print('{} {} {}'.format(target_idx, row[target_idx], cell_value))

        return cell_value

    def __get_on_air_datetime(self, row, idx_date, idx_time):

        on_air_datetime = None
        on_air_date_str = None
        date_value = row[idx_date].value
        time_value = str(row[idx_time].value)
        # print('date_value {} time_value {}'.format(date_value, time_value))
        try:
            if date_value is not None:
                if type(date_value) is not datetime:
                    on_air_date_str = datetime.strftime(on_air_datetime, '%Y/%m/%d')
                else:
                    # on_air_datetime = datetime.strptime(date_value, '%Y/%m/%d')
                    on_air_date_str = datetime.strftime(date_value, '%Y/%m/%d')
        except:
            print('except on_air_date {}'.format(on_air_datetime))

        is_time_flag = True
        if time_value is not None:
            time_list = time_value.split(':')
            if len(time_list) == 2:
                on_air_datetime = datetime.strptime('{} {}'.format(on_air_date_str, time_value), '%Y/%m/%d %H:%M')
            else:
                on_air_datetime = datetime.strptime(on_air_date_str, '%Y/%m/%d')
            # print('on_air_datetime {} '.format(on_air_datetime))
        else:
            on_air_datetime = datetime.strptime(on_air_date_str, '%Y/%m/%d')
            is_time_flag = False
            # print('on_air_datetime {} '.format(on_air_datetime))

        return on_air_datetime, is_time_flag

    def export(self):
        """
        「TV録画」のシートの主要な列以外の列に何が入っているかの確認用
        :return:
        """
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb['TV録画']
        # rows = ws['A540':'V570']
        rows = ws['A1':'V9263']
        # rows = ws['A1':'V10']
        program_list = self.program_dao.get_where_list()
        recorded_list = self.recorded_dao.get_where_list()
        with open('tv_cosnle_output.txt', 'w') as fout:
            for row_idx, row in enumerate(rows):

                if row_idx == 0:
                    continue

                if row[0].value is None:
                    print('diskNoがないためSKIP')
                    continue
                tv_data = TvRecordedData()
                program_id = self.__check_program_id(7, row)

                tv_data.channelNo = program_id // 1000
                tv_data.channelSeq = program_id % 1000
                tv_data.diskNo = row[0].value if row[0].value else ''
                tv_data.seqNo = row[1].value if row[0].value else ''
                tv_data.ripStatus = tv_data.get_rip_status(row[2].value, row[4].value)
                tv_data.onAirDate = row[3].value
                tv_data.timeFlag = False
                tv_data.timeStr = row[12].value if row[12].value else ''
                tv_data.minute = self.__get_minute(tv_data.timeStr)
                tv_data.detail = row[10].value if row[10].value else ''
                tv_data.source = 'excel'

                # '=IF(ISBLANK(D3),"",TEXT(D3,"aaa"))'
                weekday_formula = '=IF(ISBLANK(D{}),"",TEXT(D{},"aaa"))'.format(row_idx+15, row_idx+15)
                # '=IF(F146="","",VLOOKUP(F146,番組名!$A$2:$C$3311,3))'
                display_formula = '=IF(F{}="","",VLOOKUP(F{},番組名!$A$2:$C$3311,3))'.format(row_idx+15, row_idx+15)
                output_line = '{}\t{}\t{}\t{}' '\t{}\t{}{}\t{}\t\t' '{}\t{}'.format(
                    tv_data.diskNo, tv_data.seqNo, tv_data.ripStatus, tv_data.onAirDate
                    , weekday_formula, tv_data.channelNo, str(tv_data.channelSeq).zfill(3), display_formula
                    , tv_data.timeStr, tv_data.detail
                )
                try:
                    fout.write(output_line + '\n')
                except:
                    print(sys.exc_info())
                    print(output_line)


if __name__ == '__main__':
    tv_contents_register = TvContentsRegister()
    tv_contents_register.export()
